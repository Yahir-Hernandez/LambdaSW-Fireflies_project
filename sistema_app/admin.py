"""Configuración del panel de administración de Django para el festival."""

import json
import openpyxl
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment

from django.conf import settings as dj_settings
from django.db.models import Avg, Count
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib import admin, messages
from django.contrib.admin import AdminSite
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.shortcuts import render
from django.urls import path, reverse as url_reverse
from django.utils.decorators import method_decorator
from django.utils.html import format_html
from django.views.decorators.http import require_GET, require_POST

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.utils import timezone

from .models import CancellationLog, Lodging, Park, Reservation, Review, Service
from .services import ReservationCheckinService, ReservationService
from .utils import generate_qr_png


class LuciernagsAdminSite(AdminSite):
    """AdminSite personalizado: redirige al dashboard de estadísticas como página inicial
    y deshabilita el cambio de contraseña desde el panel."""

    def get_urls(self):
        custom = [
            path(
                "estadisticas/",
                self.admin_view(self.stats_view),
                name="estadisticas",
            ),
            path(
                "explorar/",
                self.admin_view(self.db_view),
                name="explorar",
            ),
        ]
        return custom + super().get_urls()

    def index(self, request, extra_context=None):
        return HttpResponseRedirect(url_reverse("admin:estadisticas"))

    def password_change(self, request, extra_context=None):
        return HttpResponseRedirect(url_reverse("admin:estadisticas"))

    def password_change_done(self, request, extra_context=None):
        return HttpResponseRedirect(url_reverse("admin:estadisticas"))

    def stats_view(self, request):
        top_parks_rating = list(
            Park.objects.filter(is_deleted=False)
            .annotate(avg_rating=Avg("reviews__rating"), review_count=Count("reviews"))
            .filter(review_count__gt=0)
            .order_by("-avg_rating")[:3]
            .values("name", "avg_rating", "review_count")
        )

        lodging_prefs = list(
            Reservation.objects
            .values("lodging__kind")
            .annotate(total=Count("id"))
            .order_by("lodging__kind")
        )

        top_parks_reservations = list(
            Park.objects.filter(is_deleted=False)
            .annotate(total=Count("reservations"))
            .order_by("-total")[:5]
            .values("name", "total")
        )

        cancelled = CancellationLog.objects.count()
        active = Reservation.objects.filter(status=Reservation.Status.ACTIVE).count()
        past = Reservation.objects.filter(status=Reservation.Status.PAST).count()
        used = Reservation.objects.filter(status=Reservation.Status.USED).count()
        total = active + cancelled + past + used
        cancellation_rate = round(cancelled / total * 100, 1) if total else 0

        total_parks = Park.objects.filter(is_deleted=False).count()
        total_users = User.objects.filter(is_active=True).count()
        total_reviews = Review.objects.count()
        avg_rating_raw = Review.objects.aggregate(avg=Avg("rating"))["avg"] or 0
        avg_rating = round(float(avg_rating_raw), 1)

        # Serialización segura de Decimals para Chart.js
        def _safe(val):
            return float(val) if isinstance(val, Decimal) else val

        top_parks_rating_safe = [
            {**r, "avg_rating": _safe(r["avg_rating"])} for r in top_parks_rating
        ]

        context = {
            **self.each_context(request),
            "title": "Estadísticas del Sistema",
            "total_parks": total_parks,
            "total_users": total_users,
            "total_reservations": total,
            "total_reviews": total_reviews,
            "avg_rating": avg_rating,
            "cancellation_rate": cancellation_rate,
            "top_parks_rating_json": json.dumps(top_parks_rating_safe),
            "lodging_prefs_json": json.dumps(lodging_prefs),
            "top_parks_reservations_json": json.dumps(top_parks_reservations),
            "status_dist_json": json.dumps([
                {"status": "Activa", "total": active},
                {"status": "Cancelada", "total": cancelled},
                {"status": "Pasada", "total": past},
                {"status": "Usada", "total": used},
            ]),
        }
        return render(request, "admin/estadisticas.html", context)

    def db_view(self, request):
        app_list = self.get_app_list(request)
        context = {
            **self.each_context(request),
            "title": "Explorar Base de Datos",
            "app_list": app_list,
        }
        return render(request, "admin/explorar.html", context)


# Reemplazar la clase del singleton admin.site en-lugar para preservar el _registry
# (modelos ya registrados por otras apps, como User de django.contrib.auth).
admin.site.__class__ = LuciernagsAdminSite


admin.site.unregister(User)


@admin.register(User)
class CustomUserAdmin(UserAdmin):
    """Editar usuario: solo contraseña y permisos; el resto es solo-lectura.

    El nombre de usuario, nombre, apellido, correo, último acceso y fecha de
    alta se muestran como texto plano al editar. La contraseña sigue siendo
    editable mediante el enlace estándar al formulario de cambio.
    """

    def get_readonly_fields(self, request, obj=None):
        # En el alta (obj is None) no se bloquea nada para poder capturar el
        # username y demás datos. La restricción aplica solo al editar.
        if obj is None:
            return ()
        return ("username", "first_name", "last_name", "email", "last_login", "date_joined")


class LodgingInline(admin.TabularInline):
    """Permite editar los hospedajes de un parque desde la pantalla del propio parque."""

    model = Lodging
    extra = 0
    fields = ("kind", "name", "capacity", "description")


@admin.register(Park)
class ParkAdmin(admin.ModelAdmin):
    """Administración de parques con borrado lógico para preservar reservas."""

    list_display = (
        "name",
        "camping_capacity",
        "has_cabins_display",
        "is_deleted",
    )
    list_filter = ("is_deleted",)
    search_fields = ("name", "address")
    filter_horizontal = ("services",)
    inlines = [LodgingInline]
    actions = ["soft_delete_selected", "restore_selected"]

    def get_queryset(self, request):
        """Incluye también los parques eliminados para poder restaurarlos."""
        return Park.objects.all()

    def delete_model(self, request, obj):
        """Aplica borrado lógico cuando el administrador elimina un parque desde el detalle."""
        obj.soft_delete()

    def delete_queryset(self, request, queryset):
        """Aplica borrado lógico a varios parques seleccionados desde el listado."""
        for park in queryset:
            park.soft_delete()

    @admin.display(description="Tiene cabañas", boolean=True)
    def has_cabins_display(self, obj):
        """Columna del listado que indica si el parque tiene al menos una cabaña."""
        return obj.has_cabins

    @admin.action(description="Eliminar lógicamente los parques seleccionados")
    def soft_delete_selected(self, request, queryset):
        """Acción masiva que marca como eliminados los parques seleccionados."""
        for park in queryset:
            park.soft_delete()
        self.message_user(
            request,
            f"{queryset.count()} parque(s) eliminados lógicamente.",
            level=messages.SUCCESS,
        )

    @admin.action(description="Restaurar los parques seleccionados")
    def restore_selected(self, request, queryset):
        """Acción masiva que vuelve a habilitar los parques previamente eliminados."""
        for park in queryset:
            park.restore()
        self.message_user(
            request,
            f"{queryset.count()} parque(s) restaurados.",
            level=messages.SUCCESS,
        )


@admin.register(Service)
class ServiceAdmin(admin.ModelAdmin):
    """Administración del catálogo de servicios que se asocian a los parques."""

    list_display = ("name",)
    search_fields = ("name",)


@admin.register(Lodging)
class LodgingAdmin(admin.ModelAdmin):
    """Administración de cabañas y parcelas de camping."""

    list_display = ("name", "park", "kind", "capacity")
    list_filter = ("kind", "park")
    search_fields = ("name", "park__name")


@admin.register(Reservation)
class ReservationAdmin(admin.ModelAdmin):
    """Administración de reservaciones.

    Las reservaciones se crean siempre desde el frontend de los usuarios.
    El panel solo permite cancelarlas, cambiar su estado y validarlas en
    la entrada mediante el escaneo del QR. Los demás campos se muestran
    como solo lectura para preservar las reglas de negocio.
    """

    list_display = (
        "id",
        "user",
        "user_email",
        "park",
        "lodging",
        "start_date",
        "duration_days",
        "people",
        "status",
        "qr_link",
    )
    list_filter = ("status", "lodging__kind", "park")
    search_fields = ("user__username", "user__email", "park__name", "lodging__name")
    actions = ["cancel_reservations", "export_as_pdf", "export_as_xlsx"]
    fieldsets = (
        ("Reservación", {
            "fields": (
                "user", "park", "lodging",
                "start_date", "end_date", "people",
                "status",
                "created_at", "checkin_token",
            ),
        }),
    )
    change_list_template = "admin/sistema_app/reservation/change_list.html"

    def has_add_permission(self, request):
        """Bloquea la creación manual desde el panel para no saltarse las validaciones."""
        return False

    def get_readonly_fields(self, request, obj=None):
        """Solo deja editable el estado; el resto se muestra como texto plano."""
        return (
            "user", "park", "lodging",
            "start_date", "end_date", "people",
            "created_at", "checkin_token",
        )

    # URLs personalizadas para el flujo de check-in por código QR.

    def get_urls(self):
        """Registra las rutas adicionales del scanner y del check-in por QR."""
        urls = super().get_urls()
        custom = [
            path(
                "scan/",
                self.admin_site.admin_view(self.scan_view),
                name="sistema_app_reservation_scan",
            ),
            path(
                "checkin/<uuid:token>/data/",
                self.admin_site.admin_view(self.checkin_data_view),
                name="sistema_app_reservation_checkin_data",
            ),
            path(
                "checkin/<uuid:token>/confirm/",
                self.admin_site.admin_view(self.checkin_confirm_view),
                name="sistema_app_reservation_checkin_confirm",
            ),
            path(
                "<int:pk>/qr.png",
                self.admin_site.admin_view(self.qr_png_view),
                name="sistema_app_reservation_qr_png",
            ),
        ]
        return custom + urls

    @admin.display(description="QR")
    def qr_link(self, obj):
        """Columna del listado con un enlace para ver o imprimir el QR de la reserva."""
        url = url_reverse(
            "admin:sistema_app_reservation_qr_png", args=[obj.pk]
        )
        return format_html('<a href="{}" target="_blank">Ver QR</a>', url)

    def scan_view(self, request):
        """Muestra la página HTML con el scanner de códigos QR."""
        context = {
            **self.admin_site.each_context(request),
            "title": "Escanear QR de check-in",
        }
        return render(request, "admin/sistema_app/reservation/scan.html", context)

    @method_decorator(require_GET)
    def checkin_data_view(self, request, token):
        """
        Devuelve los datos de la reserva identificada por el token en formato JSON.

        Lo consume el scanner del panel para mostrar el resumen de la
        reservación antes de confirmar la entrada del visitante.
        """
        try:
            reservation = (
                Reservation.objects.select_related("user", "park", "lodging")
                .get(checkin_token=token)
            )
        except Reservation.DoesNotExist:
            return JsonResponse({"error": "Reserva no encontrada."}, status=404)
        return JsonResponse({
            "id": reservation.pk,
            "token": str(reservation.checkin_token),
            "user_name": (
                reservation.user.get_full_name() or reservation.user.username
            ),
            "user_email": reservation.user.email,
            "park": reservation.park.name,
            "lodging": str(reservation.lodging),
            "start_date": reservation.start_date.isoformat(),
            "end_date": reservation.end_date.isoformat(),
            "people": reservation.people,
            "status": reservation.status,
            "status_display": reservation.get_status_display(),
            "can_check_in": reservation.status == Reservation.Status.ACTIVE,
        })

    @method_decorator(require_POST)
    def checkin_confirm_view(self, request, token):
        """Confirma el check-in y deja la reservación en estado usada."""
        try:
            reservation = Reservation.objects.get(checkin_token=token)
        except Reservation.DoesNotExist:
            return JsonResponse({"error": "Reserva no encontrada."}, status=404)
        try:
            updated = ReservationCheckinService.check_in(reservation)
        except ValidationError as exc:
            return JsonResponse({"error": "; ".join(exc.messages)}, status=400)
        return JsonResponse({
            "ok": True,
            "status": updated.status,
            "status_display": updated.get_status_display(),
        })

    @method_decorator(require_GET)
    def qr_png_view(self, request, pk):
        """
        Devuelve el código QR de la reservación como imagen PNG.

        Sirve para pruebas locales sin tener que extraer la imagen del
        correo y también para reimprimir el QR si el usuario perdió el
        correo original.
        """
        try:
            reservation = Reservation.objects.get(pk=pk)
        except Reservation.DoesNotExist:
            return HttpResponse(status=404)

        checkin_path = url_reverse(
            "admin:sistema_app_reservation_checkin_data",
            args=[reservation.checkin_token],
        )
        checkin_url = f"{dj_settings.SITE_URL.rstrip('/')}{checkin_path}"
        png = generate_qr_png(checkin_url)
        return HttpResponse(png, content_type="image/png")

    @admin.display(description="Correo")
    def user_email(self, obj):
        """Columna del listado con el correo del usuario que hizo la reservación."""
        return obj.user.email

    @admin.display(description="Duración (días)")
    def duration_days(self, obj):
        """Columna del listado con la cantidad de días que dura la reservación."""
        return (obj.end_date - obj.start_date).days

    @admin.action(description="Cancelar las reservaciones seleccionadas")
    def cancel_reservations(self, request, queryset):
        """Acción masiva que cancela las reservaciones aplicando las reglas de negocio."""
        cancelled = 0
        errors = 0
        for reservation in queryset:
            try:
                ReservationService.cancel_reservation(request.user, reservation)
                cancelled += 1
            except ValidationError:
                errors += 1
        if cancelled:
            self.message_user(
                request,
                f"{cancelled} reservación(es) canceladas.",
                level=messages.SUCCESS,
            )
        if errors:
            self.message_user(
                request,
                f"{errors} reservación(es) no pudieron cancelarse.",
                level=messages.WARNING,
            )

    @admin.action(description="Exportar reservaciones seleccionadas a PDF")
    def export_as_pdf(self, request, queryset):
        """Genera un reporte en formato PDF con las reservaciones seleccionadas."""

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_reservaciones.pdf"'

        # Documento en orientación horizontal.
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter),
            rightMargin=30, leftMargin=30,
            topMargin=30, bottomMargin=20
        )
        elements = []

        # Estilos del título y subtítulo del reporte.
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#2C3E50"),
            alignment=1,
            spaceAfter=10
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.gray,
            alignment=1,
            spaceAfter=20
        )

        elements.append(Paragraph("Reporte de Reservaciones", title_style))
        elements.append(Paragraph("Festival Internacional de las Luciérnagas 2026", subtitle_style))

        # Datos de la tabla, con encabezado en la primera fila.
        data = [['ID', 'Cliente', 'Correo', 'Parque', 'Hospedaje', 'Inicio', 'Fin', 'Pax', 'Estado']]

        for obj in queryset:
            client = obj.user.get_full_name() or obj.user.username
            lodging = obj.lodging.name if obj.lodging else 'N/A'

            data.append([
                str(obj.id),
                client,
                obj.user.email,
                obj.park.name,
                lodging,
                obj.start_date.strftime('%d/%m/%Y'),
                obj.end_date.strftime('%d/%m/%Y'),
                str(obj.people),
                obj.status
            ])

        table = Table(data, repeatRows=1)

        # Estilos visuales de la tabla.
        table_styles = TableStyle([

            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#27AE60")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#BDC3C7")),

        ])

        # Filas alternadas con color de fondo distinto.
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_styles.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#EAEDED"))
            else:
                table_styles.add('BACKGROUND', (0, i), (-1, i), colors.white)

        table.setStyle(table_styles)
        elements.append(table)

        # Pie de página con la fecha en que se generó el reporte.
        elements.append(Spacer(1, 0.2 * inch))
        date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=2)
        fecha_actual = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')
        elements.append(Paragraph(f"Reporte generado el: {fecha_actual}", date_style))

        doc.build(elements)
        return response


    @admin.action(description="Exportar reservaciones seleccionadas a Excel (XLSX)")
    def export_as_xlsx(self, request, queryset):
        """Genera un archivo XLSX con las reservaciones seleccionadas."""

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_reservaciones.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservaciones"

        headers = ['ID', 'Cliente', 'Correo', 'Parque', 'Hospedaje', 'Inicio', 'Fin', 'Pax', 'Estado']
        ws.append(headers)

        # Estilo de los encabezados de la primera fila.
        fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)

        for celda in ws[1]:
            celda.fill = fill
            celda.font = font
            celda.alignment = Alignment(horizontal="center", vertical="center")

        for obj in queryset:
            client = obj.user.get_full_name() or obj.user.username
            lodging = obj.lodging.name if obj.lodging else 'N/A'

            ws.append([
                obj.id,
                client,
                obj.user.email,
                obj.park.name,
                lodging,
                obj.start_date.strftime('%d/%m/%Y'),
                obj.end_date.strftime('%d/%m/%Y'),
                obj.people,
                obj.status
            ])

        # Ajustar el ancho de las columnas según el contenido más largo.
        for col in ws.columns:
            max_length = 0
            columna_letra = col[0].column_letter

            for celda in col:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(celda.value)
                except:
                    pass

            ws.column_dimensions[columna_letra].width = max_length + 2

        wb.save(response)
        return response


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "park", "rating", "created_at")
    list_filter = ("rating", "park")
    search_fields = ("user__username", "park__name", "comment")
    readonly_fields = ("user", "park", "reservation", "created_at")
